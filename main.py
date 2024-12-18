import tkinter as tk
import xml.etree.ElementTree as ET
import re
import statistics
import csv
import os
import json
from datetime import datetime
from tkinter import messagebox, filedialog
from tkinter import simpledialog
from tkinter import ttk, colorchooser
from collections import deque
import openpyxl
import xlwt
import xlrd
import pandas as pd
import ezodf
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from ezodf import Sheet
from odf.opendocument import OpenDocumentText
from odf.text import P
version = '(Version 0.73 FINAL BETA) '
window = tk.Tk()
window.file_extension = ''
listbox = None
counter = 1
SESSION_FILE = "session.json"

def remove_duplicates(listbox):
    unique_items = []
    for i in range(listbox.size()):
        item = listbox.get(i).split(". ")[1]
        if item not in unique_items:
            unique_items.append(item)
    
    listbox.delete(0, tk.END)
    for i, item in enumerate(unique_items, start=1):
        listbox.insert(tk.END, f"{i}. {item}")
    
    update_status(status_label, "Duplicates removed. List updated.")
    messagebox.showinfo("Duplicates Removed", "All duplicate entries have been removed.")

def validate_input(value):
    try:
        if '.' in value:
            return float(value)  # Return as float if it contains a decimal point
        return int(value)  # Return as integer otherwise
    except ValueError:
        raise ValueError(f"Invalid data: {value}")

def create_status_bar(window):
    status_label = tk.Label(window, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W)
    status_label.pack(side=tk.BOTTOM, fill=tk.X)
    return status_label

def update_status(status_label, message):
    status_label.config(text=message)

def bind_keyboard_shortcuts(window, listbox, undo_redo_manager, entry):
    window.bind("<Control-z>", lambda event: undo(undo_redo_manager, listbox))
    window.bind("<Control-y>", lambda event: redo(undo_redo_manager, listbox))
    window.bind("<Control-n>", lambda event: add_number(window, listbox, entry, undo_redo_manager, history_manager))
    window.bind("<Control-s>", lambda event: session_manager.save_session(listbox))
    window.bind("<Control-q>", lambda event: exit_file(window))

class SessionManager():
  def __init__(self, session_file=SESSION_FILE):
    self.session_file = session_file

  def save_session(self, listbox):
    data = list(listbox.get(0, tk.END))
    try:
      with open(self.session_file, "w") as file:
        json.dump(data, file)
    except Exception as e:
      messagebox.showerror("Error", f"Failed to save session: {e}")
  
  def load_session(self, listbox):
    if os.path.exists(self.session_file):
      try:
        with open(self.session_file, "r") as file:
          data = json.load(file)
        if data:
          if messagebox.askyesno("Recover Session", "Do you want to recover the last session?"):
            listbox.delete(0, tk.END)
            for i, item in enumerate(data, start=1):
              listbox.insert(tk.END, f"{i}. {item}")
      except Exception as e:
        messagebox.showerror("Error", f"Failed to load session: {e}")
  
  def clear_session(self):
    if os.path.exists(self.session_file):
      try:
        os.remove(self.session_file)
      except Exception as e:
        messagebox.showerror("Error", f"Failed to clear session: {e}")

session_manager = SessionManager()

class DataTransformer():
  def __init__(self, data):
    self.data = np.array(data)

  def min_max_scaling(self):
    min_vals = np.min(self.data, axis=0, keepdims=True)
    max_vals = np.max(self.data, axis=0, keepdims=True)
    return (self.data - min_vals) / (max_vals - min_vals)

  def z_score_normalisation(self):
    mean_vals = np.mean(self.data, axis=0, keepdims=True)
    std_vals = np.std(self.data, axis=0, keepdims=True)
    return (self.data - mean_vals) / std_vals
  
  def logarithmic_scaling(self):
    return np.log1p(self.data - np.min(self.data))
  
  def exponential_scaling(self, base=2):
    return np.power(base, self.data)
  
  def equal_width_binning(self, num_bins):
    bins = np.linspace(np.min(self.data), np.max(self.data), num_bins + 1)
    return np.digitize(self.data, bins) - 1
  
  def equal_frequency_binning(self, num_bins):
    return pd.qcut(self.data.flatten(), q=num_bins, labels=False).reshape(self.data.shape)
  
  def detect_outliers_iqr(self, factor=1.5):
    Q1 = np.percentile(self.data, 25, axis=0, keepdims=True)
    Q3 = np.percentile(self.data, 75, axis=0, keepdims=True)
    IQR = Q3 - Q1
    lower_bound = Q1 - factor * IQR
    upper_bound = Q3 + factor * IQR
    return (self.data < lower_bound) | (self.data > upper_bound)

  def detect_outliers_zscore(self, threshold=3):
    z_scores = np.abs(stats.zscore(self.data, axis=0))
    return z_scores > threshold
  
  def remove_outliers(self, method='iqr'):
    if method == 'iqr':
      outliers = self.detect_outliers_iqr()
    elif method == 'zscore':
      outliers = self.detect_outliers_zscore()
    else:
      raise ValueError("Method must be either 'iqr' or 'zscore'")
    return self.data[~outliers.any(axis=1)]
  
  def cap_outliers(self, method='iqr'):
    if method == 'iqr':
      outliers = self.detect_outliers_iqr()
      Q1 = np.percentile(self.data, 25, axis=0, keepdims=True)
      Q3 = np.percentile(self.data, 75, axis=0, keepdims=True)
      IQR = Q3 - Q1
      lower_bound = Q1 - 1.5 * IQR
      upper_bound = Q3 + 1.5 * IQR
    elif method == 'zscore':
      outliers = self.detect_outliers_zscore()
      lower_bound = np.mean(self.data, axis=0) - 3 * np.std(self.data, axis=0)
      upper_bound = np.mean(self.data, axis=0) + 3 * np.std(self.data, axis=0)
    else:
      raise ValueError("Method must be 'iqr' or 'zscore'")
    
    return np.clip(self.data, lower_bound, upper_bound)
  
  def impute_missing_values(self, method='mean'):
    if method == 'mean':
      return np.nan_to_num(self.data, nan=np.nanmean(self.data, axis=0))
    elif method == 'median':
      return np.nan_to_num(self.data, nan=np.nanmedian(self.data, axis=0))
    elif method == 'mode':
      return np.nan_to_num(self.data, nan=stats.mode(self.data, axis=0, keepdims=True)[0][0])
    else:
      raise ValueError("Method must be 'mean', 'median', or 'mode'")
  
  def create_lag(self, lag):
    return np.concatenate([np.full((lag, self.data.shape[1]), np.nan), self.data[:-lag]])
  
  def difference(self, order=1):
    return np.diff(self.data, n=order, axis=0)
  
  def rolling_statistic(self, window, statistic='mean'):
    if statistic == 'mean':
      return np.apply_along_axis(lambda m: np.convolve(m, np.ones(window), 'valid') / window, axis=0, arr=self.data)
    elif statistic == 'sum':
            return np.apply_along_axis(lambda m: np.convolve(m, np.ones(window), 'valid'), axis=0, arr=self.data)
    elif statistic == 'std':
      return np.apply_along_axis(lambda m: pd.Series(m).rolling(window=window).std().dropna().values, axis=0, arr=self.data)
    else:
      raise ValueError("Statistic must be 'mean', 'sum', or 'std'")

class Command:
  def execute(self):
    pass

  def undo(self):
    pass

class AddNumberCommand(Command):
  def __init__(self, listbox, number):
    self.listbox = listbox
    self.number = number
    self.index = None

  def execute(self):
    self.index = self.listbox.size()
    self.listbox.insert(tk.END, f"{self.index + 1}. {self.number}")

  def undo(self):
    if self.index is not None:
      self.listbox.delete(self.index)

class DeleteNumberCommand(Command):
  def __init__(self, listbox, index):
    self.listbox = listbox
    self.index = index
    self.number = None

  def execute(self):
    self.number = self.listbox.get(self.index)
    self.listbox.delete(self.index)

  def undo(self):
    if self.number:
      self.listbox.insert(self.index, self.number)

class UndoRedoManager:
  def __init__(self):
    self.undo_stack = deque()
    self.redo_stack = deque()

  def execute(self, command):
    command.execute()
    self.undo_stack.append(command)
    self.redo_stack.clear()

  def undo(self):
    if not self.undo_stack:
      return False
    command = self.undo_stack.pop()
    command.undo()
    self.redo_stack.append(command)
    return True
  
  def redo(self):
    if not self.redo_stack:
      return False
    command = self.redo_stack.pop()
    command.execute()
    self.undo_stack.append(command)
    return True

class SortCommand(Command):
  def __init__(self, listbox, undo_redo_manager, reverse=False):
    self.listbox = listbox
    self.undo_redo_manager = undo_redo_manager
    self.reverse = reverse
    self.previous_state = None
    self.sorted_numbers = None
  
  def execute(self):
    self.previous_state = list(self.listbox.get(0, tk.END))
    if not self.previous_state:
      messagebox.showerror("Error", "The list is empty! Cannot sort an empty list.", parent=window)
      return
    numbers = [float(item.split(". ")[1]) for item in self.previous_state]
    numbers.sort(reverse=self.reverse)
    self.sorted_numbers = [f"{i + 1}. {number}" for i, number in enumerate(numbers)]
    self.update_listbox(self.sorted_numbers)

  def undo(self):
    if self.previous_state:
      self.update_listbox(self.previous_state)

  def update_listbox(self, numbers):
    self.listbox.delete(0, tk.END)
    for number in numbers:
      self.listbox.insert(tk.END, number)

class FilterCommand:
  def __init__(self, listbox, filter_function, description):
      self.listbox = listbox
      self.filter_function = filter_function
      self.description = description
      self.original_numbers = [item.split(". ")[1] for item in self.listbox.get(0, tk.END)]

  def execute(self):
      filtered_numbers = list(filter(self.filter_function, self.original_numbers))
      self.listbox.delete(0, tk.END)
      for i, num in enumerate(filtered_numbers, start=1):
        self.listbox.insert(tk.END, f"{i}. {num}")

  def undo(self):
      self.listbox.delete(0, tk.END)
      for i, num in enumerate(self.original_numbers, start=1):
        self.listbox.insert(tk.END, f"{i}. {num}")

  def update_listbox(self, numbers):
      self.listbox.delete(0, tk.END)
      for number in numbers:
        self.listbox.insert(tk.END, number)

class HistoryManager:
  def __init__(self):
      self.history = []
      self.current_index = -1

  def add_action(self, action):
      self.history = self.history[:self.current_index + 1]
      self.history.append(action)
      self.current_index += 1

  def undo(self):
      if self.current_index >= 0:
          self.current_index -= 1
          return self.history[self.current_index + 1]
      return None

  def redo(self):
      if self.current_index < len(self.history) - 1:
          self.current_index += 1
          return self.history[self.current_index]
      return None

  def get_current_action(self):
      if 0 <= self.current_index < len(self.history):
          return self.history[self.current_index]
      return None

  def add_state(self, state, name=None):
    if self.current_index < len(self.history) - 1:
      self.history = self.history[:self.current_index + 1]
    
    version_name = name if name else f"Version {len(self.history) + 1}"
    self.history.append((version_name, state))
    self.current_index += 1

  def get_history(self):
    return self.history
  
  def get_current_version_name(self):
    if self.history:
      return self.history[self.current_index][0]
    return None
  
def view_history(history_manager):
  history = history_manager.get_history()
  if not history:
      messagebox.showinfo("Info", "No history available.", parent=window)
      return

  history_window = tk.Toplevel(window)
  history_window.title("History")

  listbox_history = tk.Listbox(history_window, width=50, height=10)
  listbox_history.pack()

  for name, _ in history:
    listbox_history.insert(tk.END, name)

  def restore_selected():
    selected_index = listbox_history.curselection()
    if selected_index:
      _, state_to_restore = history[selected_index[0]]
      restore_history(state_to_restore)

  restore_button = tk.Button(history_window, text="Restore Selected", command=restore_selected)
  restore_button.pack()

def display_current_version_name(window, history_manager, version_label):
  version_name = history_manager.get_current_version_name()
  if version_name:
    version_label.config(text=f"Current Version: {version_name}")
  else:
    version_label.config(text="No version available")


def save_named_version(window, listbox, history_manager):
  version_name = simpledialog.askstring("Version Name", "Enter a name for this version:", parent=window)
  if not version_name:
    messagebox.showerror("Error", "Version name cannot be empty.", parent=window)
    return
  history_manager.add_state(list(listbox.get(0, tk.END)), name=version_name)

def restore_history(state):
  listbox.delete(0, tk.END)
  for item in state:
    listbox.insert(tk.END, item)

def delete_selected_entry(listbox, undo_redo_manager):
  selected_index = listbox.curselection()
  if selected_index:
    command = DeleteNumberCommand(listbox, selected_index[0])
    undo_redo_manager.execute(command)
  else:
    messagebox.showerror("Error", "No entry selected!", parent=window)

def undo(undo_redo_manager, listbox):
  if undo_redo_manager.undo():
    update_listbox_numbers(listbox)
    update_status(status_label, f"Undo performed. List size: {listbox.size()}")
  else:
    messagebox.showinfo("Info","Nothing to undo!")

def redo(undo_redo_manager, listbox):
  if undo_redo_manager.redo():
    update_listbox_numbers(listbox)
    update_status(status_label, f"Redo performed. List size: {listbox.size()}")
  else:
    messagebox.showinfo("Info", "Nothing to redo!")

def update_listbox_numbers(listbox):
  for i in range(listbox.size()):
    item = listbox.get(i)
    number = item.split(". ")[1]
    listbox.delete(i)
    listbox.insert(i, f"{i + 1}. {number}")

def add_number(window, listbox, entry, undo_redo_manager, history_manager):
    global counter, status_label
    item = entry.get().strip()
    if not item:
        messagebox.showerror("Error", "You can't add an empty entry!", parent=window)
        return
    try:
        # Validate input
        valid_item = validate_input(item)
        
        # Add the number or valid item
        command = AddNumberCommand(listbox, str(valid_item))
        undo_redo_manager.execute(command)
        counter += 1
        history_manager.add_state(list(listbox.get(0, tk.END)))
        session_manager.save_session(listbox)
        update_status(status_label, f"Number added: {valid_item}. List size: {listbox.size()}")
    except ValueError as e:
        messagebox.showerror("Input Error", str(e), parent=window)
        entry.config(fg="red")  # Highlight the entry field in red
        window.after(2000, lambda: entry.config(fg="black"))  # Reset after 2 seconds
    finally:
        entry.delete(0, tk.END)

def clear_list(listbox, history_manager, show_message=True):
    if not listbox.get(0, tk.END) and show_message:
        messagebox.showerror("Error", "The list is already empty!")
    else:
        listbox.delete(0, tk.END)
        history_manager.add_state([])
        session_manager.clear_session()
        update_status(status_label, "List cleared.")

from odf import text, teletype
from odf.opendocument import load

def open_file(window, listbox):
  global counter
  filename = filedialog.askopenfilename(filetypes=[
    ('Excel Files (.xls, .xlsx)', '*.xls ; *.xlsx'), ('CSV Files (.csv)', '*.csv'),
    ('ODF Text Files (.odt)', '*.odt'), ('ODF Spreadsheet Files (.ods)', '*.ods'),
    ('All Files', '*.*')
  ], parent=window)
  if not filename:
    return
  clear_list(listbox, history_manager, show_message=False)
  window.file_extension = "." + filename.split(".")[-1]
  skipped_rows = []
  try:
    if filename.endswith('.xls'):
      book = xlrd.open_workbook(filename)
      sheet = book.sheet_by_index(0)
      for i in range(sheet.rows):
        try:
          value = validate_input(str(sheet.cell_value(i, 0)))
          listbox.insert(tk.END, f"{counter}.{value}")
          counter += 1
        except ValueError as e:
          skipped_rows.append((i, e))
    elif filename.endswith('.csv'):
      with open(filename, 'r') as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
          for number in row:
            try:
              value = validate_input(number)
              listbox.insert(tk.END, f"{counter}.{value}")
              counter += 1
            except ValueError as e:
              skipped_rows.append((i, e))
    elif filename.endswith('.odt'):
      textdoc = load(filename)
      allparas = textdoc.getElementsByType(text.P)
      for i, para in enumerate(allparas):
        try:
          value = validate_input(teletype.extractText(para))
          listbox.insert(tk.END, f"{counter}.{value}")
          counter += 1
        except ValueError as e:
          skipped_rows.append((i, e))
    elif filename.endswith('.ods'):
      spreadsheet = ezodf.opendoc(filename).sheets[0]
      for i, row in enumerate(spreadsheet.rows()):
        for cell in row:
          try:
            if cell.value is not None:
              value = validate_input(str(cell.value))
              listbox.insert(tk.END, f"{counter}.{value}")
              counter += 1
          except ValueError as e:
            skipped_rows.append((i, e))
    elif filename.endswith('.xlsx'):
      workbook = openpyxl.load_workbook(filename)
      sheet = workbook.active
      for i, row in enumerate(sheet.iter_cols(min_row=1, min_col=1, values_only=True)):
        for cell in row:
          try:
            if cell is not None:
              value = validate_input(str(cell))
              listbox.insert(tk.END, f"{counter}.{value}")
              counter += 1
          except ValueError as e:
            skipped_rows.append((i, e))
    if skipped_rows:
      messagebox.showwarning(
        "Warning",
        f"Some rows were skipped due to invalid data:\n" +
        "\n".join([f"Row {row}: {error}" for row, error in skipped_rows])
      )
    update_status(f"File loaded: {os.path.basename(filename)}. List size: {listbox.size()}")
  except Exception as e:
    messagebox.showerror("Error", f"Failed to open file: {e}", parent=window)

def add_all_numbers(window, listbox, history_manager, version_label):
  global counter
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 2:
    messagebox.showerror("Error",
                         "There must be numbers to add!",
                         parent=window)
    return
  total = sum(int(number.split(". ")[1]) for number in numbers)
  clear_list(listbox, history_manager)
  listbox.insert(tk.END, f"{counter}. {total}")
  version_name = simpledialog.askstring("Version Name", "Please enter a name for this version:", parent=window)
  if not version_name:
    messagebox.showerror("Error", "Version name cannot be empty.", parent=window)
    return
  history_manager.add_state(list(listbox.get(0, tk.END)), name=version_name)

  display_current_version_name(window, history_manager, version_label)

def subtract_numbers(window, listbox, history_manager, version_label):
    global counter
    numbers = listbox.get(0, tk.END)
    if len(numbers) < 2:
      messagebox.showerror("Error",
                           "There must be numbers to subtract!",
                           parent=window)
      return
    total = float(numbers[0].split(". ")[1]) - sum(
      float(number.split(". ")[1]) for number in numbers[1:])
    clear_list(listbox, history_manager)
    listbox.insert(tk.END, f"{counter}. {total}")
    version_name = simpledialog.askstring("Version Name", "Please enter a name for this version:", parent=window)
    if not version_name:
      messagebox.showerror("Error", "Version name cannot be empty.", parent=window)
      return
    history_manager.add_state(list(listbox.get(0, tk.END)), name=version_name)

    display_current_version_name(window, history_manager, version_label)


def multiply_all_numbers(window, listbox, history_manager, version_label):
  global counter
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 2:
    messagebox.showerror("Error",
                         "There must be numbers to multiply!",
                         parent=window)
    return
  total = 1
  for number in numbers:
    total *= int(number.split(". ")[1])
  clear_list(listbox, history_manager)
  listbox.insert(tk.END, f"{counter}. {total}")
  version_name = simpledialog.askstring("Version Name", "Please enter a name for this version:", parent=window)
  if not version_name:
    messagebox.showerror("Error", "Version name cannot be empty.", parent=window)
    return
  history_manager.add_state(list(listbox.get(0, tk.END)), name=version_name)

  display_current_version_name(window, history_manager, version_label)

def divide_all_numbers(window, listbox, history_manager, version_label):
  global counter
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 2:
    messagebox.showerror("Error",
                         "There must be at least two numbers to divide!",
                         parent=window)
    return
  total = int(numbers[0].split(". ")[1])
  for number in numbers[1:]:
    if int(number.split(". ")[1]) == 0:
      messagebox.showerror("Error", "Cannot divide by zero!", parent=window)
      return
    total /= int(number.split(". ")[1])
  clear_list(listbox, history_manager)
  listbox.insert(tk.END, f"{counter}. {total}")
  version_name = simpledialog.askstring("Version Name", "Please enter a name for this version:", parent=window)
  if not version_name:
    messagebox.showerror("Error", "Version name cannot be empty.", parent=window)
    return
  history_manager.add_state(list(listbox.get(0, tk.END)), name=version_name)

  display_current_version_name(window, history_manager, version_label)

def square_all_numbers(window, listbox, history_manager, version_label):
  global counter
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 1:
    messagebox.showerror("Error",
                         "There must be at least one number to square!",
                         parent=window)
    return
  clear_list(listbox, history_manager)
  for number in numbers:
    square_number = float(number.split(". ")[1])**2
    listbox.insert(tk.END, f"{counter}. {square_number}")
    version_name = simpledialog.askstring("Version Name", "Please enter a name for this version:", parent=window)
    if not version_name:
      messagebox.showerror("Error", "Version name cannot be empty.", parent=window)
      return
    history_manager.add_state(list(listbox.get(0, tk.END)), name=version_name)

    display_current_version_name(window, history_manager, version_label)

algebra_dict = {}

def numeral_system_conversions(listbox, history_manager):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error", "There are no entries in the number list!", parent=window)
    return
  for item in numbers:
    global counter
    number = int(item.split(". ")[1])
    binary_number = "{0:b}".format(number)
    octal_number = "{0:o}".format(number)
    hexadecimal_number = "{0:x}".format(number)
    converted_item = f"Binary: {binary_number}"
    listbox.insert(tk.END, f"{counter}. {converted_item}")
    counter += 1
    history_manager.add_state(list(listbox.get(0, tk.END)))
    converted_item = f"Octal: {octal_number}"
    listbox.insert(tk.END, f"{counter}. {converted_item}")
    counter += 1
    history_manager.add_state(list(listbox.get(0, tk.END)))
    converted_item = f"Hexadecimal: {hexadecimal_number}"
    listbox.insert(tk.END, f"{counter}. {converted_item}")
    counter += 1
    history_manager.add_state(list(listbox.get(0, tk.END)))

def define_algebraic_letter(window):
  letter_window = tk.Toplevel(window)
  letter_window.title("Define Algebraic Letter")
  tk.Label(letter_window, text="Algebraic Letter:").pack()
  letter_entry = tk.Entry(letter_window)
  letter_entry.pack()
  tk.Label(letter_window, text="Equivalent Number:").pack()
  number_entry = tk.Entry(letter_window)
  number_entry.pack()

  def save_algebraic_letter():
    global algebra_dict
    algebra_dict[letter_entry.get()] = int(number_entry.get())
    letter_window.destroy()

  tk.Button(letter_window, text="Save", command=save_algebraic_letter).pack()

def convert_algebra(window, listbox):
  global counter
  counter = 1
  items = listbox.get(0, tk.END)
  converted_items = []
  for item in items:
    if item in algebra_dict:
      converted_items.append(str(algebra_dict[item]))
    else:
      converted_items.append(item)
    listbox.delete(0, tk.END)
  for item in converted_items:
    listbox.insert(tk.END, f"{counter}. {item}")
    counter += 1
  if set(items) == set(converted_items):
    messagebox.showerror("Error",
                         "You can't convert non-defined algebraic numbers!",
                         parent=window)
  
def change_file_extension(window, extension):
  window.file_extension = extension

about_window = None

def about(window):
  global about_window
  if about_window is not None:
    try:
      about_window.lift()
    except tk.TclError:
      about_window = None
  if about_window is None:
    about_window = tk.Toplevel(window)
    about_window.protocol("WM_DELETE_WINDOW", close_about)
  title_label = tk.Label(about_window, text="About Number List:")
  title_label.pack()
  update_label = tk.Label(about_window, text="The 'No More Errors' Update")
  update_label.pack()
  version_label = tk.Label(about_window, text="Version 0.73 FINAL BETA")
  version_label.pack()
  contributor_label = tk.Label(about_window, text="Contributors:")
  contributor_label.pack()
  contributor_label2 = tk.Label(about_window, text="Tay Rake 2023 - 2024")
  contributor_label2.pack()

def change_theme(theme):
  if theme == "light":
      listbox.config(bg="white", fg="black")
  elif theme == "dark":
      listbox.config(bg="black", fg="white")

def close_about():
  global about_window
  about_window.destroy()

about_window = None
  
def exit_file(window):
    window.quit()

def report_bug(window):
  messagebox.showwarning("Telemetry","Some telemetry data is required so developer(s) can fix bugs. Such telemetry data is the application version. By closing this message, you agree to the telemetry data collection." , parent=window)
  bug_window = tk.Toplevel(window)
  bug_window.title("Report a Bug")
  bug_label = tk.Label(bug_window, text="Describe the bug:")
  bug_label.pack()
  bug_entry = tk.Text(bug_window, height=5, width=40)
  bug_entry.pack()
  save_button = tk.Button(bug_window, text="Save Report", command=lambda: save_bug_report(bug_entry.get("1.0", tk.END)))
  save_button.pack()

def save_bug_report(report):
  global version
  with open('bugs.txt', 'a') as file:
      file.write(version + report + "\n")
      messagebox.showinfo("Bug Report", "Bug reported successfully!")

def create_graph(window, listbox):
    numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
    if len(numbers) < 2:
        messagebox.showerror("Error", "You need at least two numbers to create a graph!", parent=window)
        return

    graph_window = tk.Toplevel(window)
    graph_window.title("Number List Graph")

    fig, ax = plt.subplots(figsize=(8, 6))
    ax.plot(range(1, len(numbers) + 1), numbers, marker='o')
    ax.set_xlabel('Index')
    ax.set_ylabel('Value')
    ax.set_title('Number List Graph')

    canvas = FigureCanvasTkAgg(fig, master=graph_window)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack()

    def save_graph():
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG files", "*.png"), ("All files", "*.*")])
        if file_path:
            fig.savefig(file_path)
            messagebox.showinfo("Success", f"Graph saved as {file_path}", parent=graph_window)

    save_button = tk.Button(graph_window, text="Save Graph", command=save_graph)
    save_button.pack(pady=10)
  
def create_advanced_graph(window, listbox):
    numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
    if len(numbers) < 2:
        messagebox.showerror("Error", "You need at least two numbers to create a graph!", parent=window)
        return

    graph_window = tk.Toplevel(window)
    graph_window.title("Advanced Graph")

    main_frame = ttk.Frame(graph_window)
    main_frame.pack(fill='both', expand=True, padx=10, pady=10)
    control_frame = ttk.Frame(main_frame)
    control_frame.pack(fill='x', pady=(0, 10))

    graph_types = ["Line", "Bar", "Scatter", "Pie"]
    graph_type_var = tk.StringVar(value="Line")
    ttk.Label(control_frame, text="Graph Type:").pack(side='left', padx=(0, 5))
    type_combo = ttk.Combobox(control_frame, textvariable=graph_type_var, values=graph_types, width=10)
    type_combo.pack(side='left', padx=(0, 10))

    color_var = tk.StringVar(value="#1f77b4")
    ttk.Label(control_frame, text="Graph Color:").pack(side='left', padx=(0, 5))
    color_button = ttk.Button(control_frame, text="Choose Color", width=15)
    color_button.pack(side='left', padx=(0, 10))

    title_var = tk.StringVar(value="Number List Graph")
    ttk.Label(control_frame, text="Graph Title:").pack(side='left', padx=(0, 5))
    title_entry = ttk.Entry(control_frame, textvariable=title_var, width=20)
    title_entry.pack(side='left')

    graph_frame = ttk.Frame(main_frame)
    graph_frame.pack(fill='both', expand=True)

    fig, ax = plt.subplots(figsize=(8, 6))
    canvas = FigureCanvasTkAgg(fig, master=graph_frame)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack(fill='both', expand=True)

    def choose_color():
        color = colorchooser.askcolor(color_var.get())[1]
        if color:
            color_var.set(color)
        update_graph()

    color_button.config(command=choose_color)

    def update_graph(*args):
        ax.clear()
        graph_type = graph_type_var.get()
        color = color_var.get()
        title = title_var.get()

        if graph_type == "Line":
            ax.plot(range(1, len(numbers) + 1), numbers, color=color, marker='o')
        elif graph_type == "Bar":
            ax.bar(range(1, len(numbers) + 1), numbers, color=color)
        elif graph_type == "Scatter":
            ax.scatter(range(1, len(numbers) + 1), numbers, color=color)
        elif graph_type == "Pie":
            ax.pie(numbers, labels=[f"Item {i+1}" for i in range(len(numbers))], autopct='%1.1f%%', colors=[color])

        ax.set_title(title)
        if graph_type != "Pie":
            ax.set_xlabel('Index')
            ax.set_ylabel('Value')
        
        fig.tight_layout()
        canvas.draw()

    graph_type_var.trace_add("write", update_graph)
    title_var.trace_add("write", update_graph)

    def save_graph():
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG files", "*.png"), ("All files", "*.*")])
        if file_path:
            fig.savefig(file_path)
            messagebox.showinfo("Success", f"Graph saved as {file_path}", parent=graph_window)

    button_frame = ttk.Frame(main_frame)
    button_frame.pack(fill='x', pady=(10, 0))

    update_button = ttk.Button(button_frame, text="Update Graph", command=update_graph)
    update_button.pack(side='left', padx=4)

    save_button = ttk.Button(button_frame, text="Save Graph", command=save_graph)
    save_button.pack(side='left', padx=5)

    update_graph()

def sort_numbers_ascending(window, listbox, undo_redo_manager, history_manager):
  command = SortCommand(listbox, undo_redo_manager, reverse=False)
  undo_redo_manager.execute(command)
  history_manager.add_state(list(listbox.get(0, tk.END)))
  update_status(status_label, f"List sorted in ascending order. List size: {listbox.size()}")

def sort_numbers_descending(window, listbox, undo_redo_manager, history_manager):
  command = SortCommand(listbox, undo_redo_manager, reverse=True)
  undo_redo_manager.execute(command)
  history_manager.add_state(list(listbox.get(0, tk.END)))
  update_status(status_label, f"List sorted in descending order. List size: {listbox.size()}")

def update_listbox_with_numbers(listbox, numbers):
  listbox.delete(0, tk.END)
  for i, number in enumerate(numbers, 1):
    listbox.insert(tk.END, f"{i}, {number}")

def filter_even_numbers(window, listbox, undo_redo_manager, history_manager):
    def even_filter(num):
        return float(num) % 2 == 0
    command = FilterCommand(listbox, even_filter, "Filter Even Numbers")
    undo_redo_manager.execute(command)
    history_manager.add_action("Filtered even numbers")
    update_status(status_label, f"Filter applied: Even numbers. List size: {listbox.size()}")

def filter_odd_numbers(window, listbox, undo_redo_manager, history_manager):
    def odd_filter(num):
        return float(num) % 2 != 0
    command = FilterCommand(listbox, odd_filter, "Filter Odd Numbers")
    undo_redo_manager.execute(command)
    history_manager.add_action("Filtered odd numbers")

def filter_custom_range(window, listbox, undo_redo_manager, history_manager):
    def apply_filter():
        try:
            min_val = float(min_entry.get())
            max_val = float(max_entry.get())
            def range_filter(num):
                return min_val <= float(num) <= max_val
            command = FilterCommand(listbox, range_filter, f"Filter Range {min_val} to {max_val}")
            undo_redo_manager.execute(command)
            history_manager.add_action(f"Filtered range {min_val} to {max_val}")
            filter_window.destroy()
        except ValueError:
            messagebox.showerror("Error", "Please enter valid numbers for the range.")

    filter_window = tk.Toplevel(window)
    filter_window.title("Filter Custom Range")

    tk.Label(filter_window, text="Minimum value:").pack()
    min_entry = tk.Entry(filter_window)
    min_entry.pack()

    tk.Label(filter_window, text="Maximum value:").pack()
    max_entry = tk.Entry(filter_window)
    max_entry.pack()

    tk.Button(filter_window, text="Apply Filter", command=apply_filter).pack()

def export_to_csv(window, listbox):
    numbers = listbox.get(0, tk.END)
    if not numbers:
        messagebox.showerror("Error", "The list is empty! Cannot export.", parent=window)
        return
    
    file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")], parent=window)
    if not file_path:
        return
    
    with open(file_path, mode='w', newline="") as file:
        writer = csv.writer(file)
        writer.writerow([number.split(". ")[1] for number in numbers])
    
    update_status(status_label, f"List exported to CSV: {os.path.basename(file_path)}.")
    messagebox.showinfo("Success", f"List exported to {file_path}", parent=window)

def export_to_excel(window, listbox):
    numbers = listbox.get(0, tk.END)
    if not numbers:
        messagebox.showerror("Error", "The list is empty! Cannot export.", parent=window)
        return

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], parent=window)
    if not file_path:
        return

    data = [json.loads(item.split(". ")[1]) if "[" in item else float(item.split(". ")[1]) for item in numbers]
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    if isinstance(data[0], list):
        for row in data:
            sheet.append(row)
    else:
        for value in data:
            sheet.append([value])

    workbook.save(file_path)
    messagebox.showinfo("Success", f"Data exported to {file_path}", parent=window)

def export_to_json(window, listbox):
    numbers = listbox.get(0, tk.END)
    if not numbers:
        messagebox.showerror("Error", "The list is empty! Cannot export.", parent=window)
        return

    file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")], parent=window)
    if not file_path:
        return

    data = [json.loads(item.split(". ")[1]) if "[" in item else float(item.split(". ")[1]) for item in numbers]
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)
    messagebox.showinfo("Success", f"Data exported to {file_path}", parent=window)

def export_to_ods(window, listbox):
    numbers = listbox.get(0, tk.END)
    if not numbers:
        messagebox.showerror("Error", "The list is empty! Cannot export.", parent=window)
        return

    file_path = filedialog.asksaveasfilename(defaultextension=".ods", filetypes=[("ODS files", "*.ods")], parent=window)
    if not file_path:
        return

    data = [json.loads(item.split(". ")[1]) if "[" in item else float(item.split(". ")[1]) for item in numbers]
    spreadsheet = ezodf.newdoc(doctype="ods", filename=file_path)
    sheet = ezodf.Sheet("Sheet1", size=(len(data), len(data[0]) if isinstance(data[0], list) else 1))
    spreadsheet.sheets.append(sheet)

    for i, row in enumerate(data):
        if isinstance(row, list):
            for j, value in enumerate(row):
                sheet[i, j].set_value(value)
        else:
            sheet[i, 0].set_value(row)

    spreadsheet.save()
    messagebox.showinfo("Success", f"Data exported to {file_path}", parent=window)

def copy_to_clipboard(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error", "The list is empty! Cannot copy an empty list.", parent=window)
    return
    
  data = "\n".join([number.split(". ")[1] for number in numbers])
  window.clipboard_clear()
  window.clipboard_append(data)
  messagebox.showinfo("Copy", "List copied to clipboard!", parent=window)

def share_via_email(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error", "The list is empty! Cannot share an empty list.", parent=window)
    return
  
  data = "\n".join([number.split(". ")[1] for number in numbers])
  subject = "My Number List"
  body = f"Here is my number list:\n\n{data}"
  body = body.replace("\n", "%0D%0A")
  mailto_link = f"mailto:?subject={subject}&body={body}"

  window.clipboard_clear()
  window.clipboard_append(mailto_link)
  messagebox.showinfo("Share", "Mailto link copied to clipboard! Paste it in your mail client.", parent=window)

def calculate_mean(listbox):
  numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
  if not numbers:
    messagebox.showerror("Error", "The list is empty! Cannot calculate statistics.", parent=window)
    return
  mean_value = statistics.mean(numbers)
  messagebox.showinfo("Mean", f"The mean of the list is: {mean_value}")

def calculate_median(listbox):
  numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
  if not numbers:
    messagebox.showerror("Error", "The list is empty! Cannot calculate statistics.", parent=window)
    return
  median_value = statistics.median(numbers)
  messagebox.showinfo("Median", f"The median of the list is: {median_value}")

def calculate_mode(listbox):
  numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
  if not numbers:
    messagebox.showerror("Error", "The list is empty! Cannot calculate statistics.", parent=window)
    return
  try:
    mode_value = statistics.mode(numbers)
    messagebox.showinfo("Mode", f"The mode of the list is: {mode_value}")
  except statistics.StatisticsError:
    messagebox.showinfo("Mode", "No unique mode found in the list.")

def calculate_variance(listbox):
  numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
  if not numbers:
      messagebox.showerror("Error", "The list is empty! Cannot calculate statistics.", parent=window)
      return
  variance_value = statistics.variance(numbers)
  messagebox.showinfo("Variance", f"The variance of the list is: {variance_value}")

def calculate_standard_deviation(listbox):
  numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
  if not numbers:
      messagebox.showerror("Error", "The list is empty! Cannot calculate statistics.", parent=window)
      return
  stddev_value = statistics.stdev(numbers)
  messagebox.showinfo("Standard Deviation", f"The standard deviation of the list is: {stddev_value}")

def apply_transformation(window, listbox, transformation, **kwargs):
    numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
    transformer = DataTransformer(numbers)
    
    try:
        if transformation == 'min_max':
            result = transformer.min_max_scaling()
        elif transformation == 'z_score':
            result = transformer.z_score_normalisation()
        elif transformation == 'log':
            result = transformer.logarithmic_scaling()
        elif transformation == 'exp':
            result = transformer.exponential_scaling()
        elif transformation == 'equal_width_bin':
            num_bins = kwargs.get('num_bins', 5)
            result = transformer.equal_width_binning(num_bins)
        elif transformation == 'equal_freq_bin':
            num_bins = kwargs.get('num_bins', 5)
            result = transformer.equal_frequency_binning(num_bins)
        elif transformation == 'remove_outliers':
            method = kwargs.get('method', 'iqr')
            result = transformer.remove_outliers(method)
        elif transformation == 'cap_outliers':
            method = kwargs.get('method', 'iqr')
            result = transformer.cap_outliers(method)
        elif transformation == 'impute_missing':
            method = kwargs.get('method', 'mean')
            result = transformer.impute_missing_values(method)
        elif transformation == 'lag':
            lag = kwargs.get('lag', 1)
            result = transformer.create_lag(lag)
        elif transformation == 'difference':
            order = kwargs.get('order', 1)
            result = transformer.difference(order)
        elif transformation == 'rolling':
            window = kwargs.get('window', 3)
            statistic = kwargs.get('statistic', 'mean')
            result = transformer.rolling_statistic(window, statistic)
        else:
            raise ValueError(f"Unknown transformation: {transformation}")
        
        listbox.delete(0, tk.END)
        for i, value in enumerate(result, start=1):
            listbox.insert(tk.END, f"{i}. {value}")
        
        messagebox.showinfo("Transformation Applied", f"The {transformation} transformation has been applied successfully.", parent=window)
    except Exception as e:
        messagebox.showerror("Error", str(e), parent=window)

def create_transformation_window(window, listbox, transformation):
  trans_window = tk.Toplevel(window)
  trans_window.title(f"Apply {transformation.capitalize()} Transformation")

  if transformation in ['equal_width_bin', 'equal_freq_bin']:
    ttk.Label(trans_window, text="Number of bins:").pack(pady=5)
    num_bins_entry = ttk.Entry(trans_window)
    num_bins_entry.pack(pady=5)
    num_bins_entry.insert(0, "5")

    ttk.Button(trans_window, text="Apply",
               command=lambda: apply_transformation(window, listbox, transformation,
                                                    num_bins=int(num_bins_entry.get()))).pack(pady=10)

  elif transformation in ['remove_outliers', 'cap_outliers']:
        method_var = tk.StringVar(value="iqr")
        ttk.Radiobutton(trans_window, text="IQR Method", variable=method_var, value="iqr").pack()
        ttk.Radiobutton(trans_window, text="Z-Score Method", variable=method_var, value="zscore").pack()
        
        ttk.Button(trans_window, text="Apply", 
                   command=lambda: apply_transformation(window, listbox, transformation, 
                                                        method=method_var.get())).pack(pady=10)
  
  elif transformation == 'impute_missing':
        method_var = tk.StringVar(value="mean")
        ttk.Radiobutton(trans_window, text="Mean", variable=method_var, value="mean").pack()
        ttk.Radiobutton(trans_window, text="Median", variable=method_var, value="median").pack()
        ttk.Radiobutton(trans_window, text="Mode", variable=method_var, value="mode").pack()
        
        ttk.Button(trans_window, text="Apply", 
                   command=lambda: apply_transformation(window, listbox, transformation, 
                                                        method=method_var.get())).pack(pady=10)
  
  elif transformation == 'lag':
        ttk.Label(trans_window, text="Lag value:").pack(pady=5)
        lag_entry = ttk.Entry(trans_window)
        lag_entry.pack(pady=5)
        lag_entry.insert(0, "1")
        
        ttk.Button(trans_window, text="Apply", 
                   command=lambda: apply_transformation(window, listbox, transformation, 
                                                        lag=int(lag_entry.get()))).pack(pady=10)
  elif transformation == 'difference':
    ttk.Label(trans_window, text="Difference order:").pack(pady=5)
    order_entry = ttk.Entry(trans_window)
    order_entry.pack(pady=5)
    order_entry.insert(0, "1")

    ttk.Button(trans_window, text="Apply", 
                   command=lambda: apply_transformation(window, listbox, transformation, 
                                                        order=int(order_entry.get()))).pack(pady=10)
  
  elif transformation == 'rolling':
        ttk.Label(trans_window, text="Window size:").pack(pady=5)
        window_entry = ttk.Entry(trans_window)
        window_entry.pack(pady=5)
        window_entry.insert(0, "3")
        
        statistic_var = tk.StringVar(value="mean")
        ttk.Radiobutton(trans_window, text="Mean", variable=statistic_var, value="mean").pack()
        ttk.Radiobutton(trans_window, text="Sum", variable=statistic_var, value="sum").pack()
        ttk.Radiobutton(trans_window, text="Standard Deviation", variable=statistic_var, value="std").pack()
        
        ttk.Button(trans_window, text="Apply", 
                   command=lambda: apply_transformation(window, listbox, transformation, 
                                                        window=int(window_entry.get()),
                                                        statistic=statistic_var.get())).pack(pady=10)
    
  else:
      ttk.Button(trans_window, text="Apply", 
                  command=lambda: apply_transformation(window, listbox, transformation)).pack(pady=10)

def create_histogram(window, listbox):
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

    numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
    if not numbers:
        messagebox.showerror("Error", "The list is empty! Cannot create a histogram.", parent=window)
        return

    hist_window = tk.Toplevel(window)
    hist_window.title("Histogram")

    fig, ax = plt.subplots(figsize=(8, 6))
    ax.hist(numbers, bins='auto', color='#1f77b4', edgecolor='black')
    ax.set_xlabel('Value')
    ax.set_ylabel('Frequency')
    ax.set_title('Histogram')

    canvas = FigureCanvasTkAgg(fig, master=hist_window)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack()

    def save_histogram():
        file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                 filetypes=[("PNG files", "*.png"), ("All files", "*.*")])
        if file_path:
            fig.savefig(file_path)
            messagebox.showinfo("Success", f"Histogram saved as {file_path}", parent=hist_window)

    save_button = tk.Button(hist_window, text="Save Histogram", command=save_histogram)
    save_button.pack(pady=10)

def create_box_plot(window, listbox):
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

    numbers = [float(item.split(". ")[1]) for item in listbox.get(0, tk.END)]
    if not numbers:
        messagebox.showerror("Error", "The list is empty! Cannot create a box plot.", parent=window)
        return

    box_window = tk.Toplevel(window)
    box_window.title("Box Plot")

    fig, ax = plt.subplots(figsize=(8, 6))
    ax.boxplot(numbers, vert=True, patch_artist=True)
    ax.set_ylabel('Value')
    ax.set_title('Box Plot')

    canvas = FigureCanvasTkAgg(fig, master=box_window)
    canvas_widget = canvas.get_tk_widget()
    canvas_widget.pack()

    def save_box_plot():
        file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                 filetypes=[("PNG files", "*.png"), ("All files", "*.*")])
        if file_path:
            fig.savefig(file_path)
            messagebox.showinfo("Success", f"Box plot saved as {file_path}", parent=box_window)

    save_button = tk.Button(box_window, text="Save Box Plot", command=save_box_plot)
    save_button.pack(pady=10)

def create_new_window():
  global status_label
  global counter
  window = tk.Tk()
  global listbox
  window.title("Number List")
  undo_redo_manager = UndoRedoManager()
  undo_button = tk.Button(window, text="Undo", command=lambda: undo(undo_redo_manager, listbox))
  undo_button.pack()
  redo_button = tk.Button(window, text="Redo", command=lambda: redo(undo_redo_manager, listbox))
  redo_button.pack()
  delete_button = tk.Button(window,
                            text="Delete Selected Entry",
                            command=lambda: delete_selected_entry(listbox, undo, undo_redo_manager))
  delete_button.pack()
  menubar = tk.Menu(window)
  window.config(menu=menubar)
  file_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="File", menu=file_menu)
  file_menu.add_command(label="New", command=lambda: create_new_window())
  file_menu.add_command(label="Import",
                        command=lambda: open_file(window, listbox))
  file_menu.add_command(label="Exit", command=lambda: exit_file(window))
  export_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Export", menu=export_menu)
  export_menu.add_command(label="Export to CSV", command=lambda: export_to_csv(window, listbox))
  export_menu.add_command(label="Export to JSON", command=lambda: export_to_json(window, listbox))
  export_menu.add_command(label="Export to Excel", command=lambda: export_to_excel(window, listbox))
  export_menu.add_command(label="Export to ODS", command=lambda: export_to_ods(window, listbox))
  edit_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Edit", menu=edit_menu)
  edit_menu.add_command(label="Undo", command=lambda: undo(undo_redo_manager, listbox))
  edit_menu.add_command(label="Redo", command=lambda: redo(undo_redo_manager, listbox))
  edit_menu.add_command(label="Remove Duplicates", command=lambda: remove_duplicates(listbox))
  help_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Help", menu=help_menu)
  help_menu.add_command(label="About", command=lambda: about(window))
  math_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Calculate", menu=math_menu)
  math_menu.add_command(label="Add",
                        command=lambda: add_all_numbers(window, listbox))
  math_menu.add_command(label="Subtract",
                        command=lambda: subtract_numbers(window, listbox))
  math_menu.add_command(label="Multiply",
                        command=lambda: multiply_all_numbers(window, listbox))
  math_menu.add_command(label="Divide",
                        command=lambda: divide_all_numbers(window, listbox))
  math_menu.add_command(label="Square",
                        command=lambda: square_all_numbers(window, listbox))
  math_menu.add_command(label="Define Algebraic Letter",
                        command=lambda: define_algebraic_letter(window))
  more_algebra_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="More Algebra...", menu=more_algebra_menu)
  more_algebra_menu.add_command(
    label="Convert Algebra", command=lambda: convert_algebra(window, listbox))
  math_menu.add_command(label="Numeral System Conversions", command=lambda: numeral_system_conversions(listbox))
  sort_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="Sort", menu=sort_menu)
  sort_menu.add_command(label="Ascending", command=lambda: sort_numbers_ascending(window, listbox, undo_redo_manager))
  sort_menu.add_command(label="Descending", command=lambda: sort_numbers_descending(window, listbox, undo_redo_manager))
  filter_menu = tk.Menu(math_menu, tearoff=0)
  filter_menu.add_command(label="Even Numbers", 
                            command=lambda: filter_even_numbers(window, listbox, undo_redo_manager, history_manager))
  filter_menu.add_command(label="Odd Numbers", 
                            command=lambda: filter_odd_numbers(window, listbox, undo_redo_manager, history_manager))
  filter_menu.add_command(label="Custom Range", 
                            command=lambda: filter_custom_range(window, listbox, undo_redo_manager, history_manager))
  graph_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Graph", menu=graph_menu)
  graph_menu.add_command(label="Create Graph", command=lambda: create_graph(window, listbox))
  graph_menu.add_command(label="Create Advanced Graph", command=lambda: create_advanced_graph(window, listbox))
  graph_menu.add_command(label="Create Histogram", command=lambda: create_histogram(window, listbox))
  graph_menu.add_command(label="Create Box Plot", command=lambda: create_box_plot(window, listbox))
  history_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="History", menu=history_menu)
  history_menu.add_cascade(label="View History", command=lambda: view_history(history_manager))
  history_menu.add_cascade(label="Undo to Previous State", command=lambda: restore_history(history_manager.undo()))
  history_menu.add_cascade(label="Redo to Next State", command=lambda: restore_history(history_manager.redo()))
  stats_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Statistics", menu=stats_menu)
  stats_menu.add_command(label="Mean", command=lambda: calculate_mean(listbox))
  stats_menu.add_command(label="Median", command=lambda: calculate_median(listbox))
  stats_menu.add_command(label="Mode", command=lambda: calculate_mode(listbox))
  stats_menu.add_command(label="Variance", command=lambda: calculate_variance(listbox))
  stats_menu.add_command(label="Standard Deviation", command=lambda: calculate_standard_deviation(listbox))
  transform_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Transform", menu=transform_menu)
  normalize_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Normalize", menu=normalize_menu)
  normalize_menu.add_command(label="Min-Max Scaling", 
                               command=lambda: create_transformation_window(window, listbox, 'min_max'))
  normalize_menu.add_command(label="Z-Score Normalization", 
                               command=lambda: create_transformation_window(window, listbox, 'z_score'))
  scale_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Scale", menu=scale_menu)
  scale_menu.add_command(label="Logarithmic Scaling", 
                           command=lambda: create_transformation_window(window, listbox, 'log'))
  scale_menu.add_command(label="Exponential Scaling", 
                           command=lambda: create_transformation_window(window, listbox, 'exp'))
  bin_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Binning", menu=bin_menu)
  bin_menu.add_command(label="Equal-width Binning", 
                         command=lambda: create_transformation_window(window, listbox, 'equal_width_bin'))
  bin_menu.add_command(label="Equal-frequency Binning", 
                         command=lambda: create_transformation_window(window, listbox, 'equal_freq_bin'))
  outlier_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Outlier Handling", menu=outlier_menu)
  outlier_menu.add_command(label="Remove Outliers", 
                             command=lambda: create_transformation_window(window, listbox, 'remove_outliers'))
  outlier_menu.add_command(label="Cap Outliers", 
                             command=lambda: create_transformation_window(window, listbox, 'cap_outliers'))
  transform_menu.add_command(label="Impute Missing Values", 
                               command=lambda: create_transformation_window(window, listbox, 'impute_missing'))
  time_series_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Time Series", menu=time_series_menu)
  time_series_menu.add_command(label="Create Lag", 
                                 command=lambda: create_transformation_window(window, listbox, 'lag'))
  time_series_menu.add_command(label="Difference", 
                                 command=lambda: create_transformation_window(window, listbox, 'difference'))
  time_series_menu.add_command(label="Rolling Statistics", 
                                 command=lambda: create_transformation_window(window, listbox, 'rolling'))
  listbox = tk.Listbox(window)
  listbox.pack()
  entry = tk.Entry(window)
  entry.pack()
  button_add = tk.Button(window, text="Add number", 
                         command=lambda: add_number(window, listbox, entry, undo_redo_manager, history_manager))
  button_add.pack()
  button_clear = tk.Button(window,
                           text="Clear List",
                           command=lambda: clear_list(listbox))
  button_clear.pack()
  status_label = create_status_bar(window)
  session_manager.load_session(listbox)
  update_status(status_label, f"List loaded. Current size: {listbox.size()}")
  bind_keyboard_shortcuts(window, listbox, undo_redo_manager, entry)
  theme_menu = tk.Menu(menubar, tearoff=0)
  theme_menu.add_command(label="Light Theme", command=lambda: change_theme("light"))
  theme_menu.add_command(label="Dark Theme", command=lambda: change_theme("dark"))
  menubar.add_cascade(label="Themes", menu=theme_menu)
  share_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Share", menu=share_menu)
  share_menu.add_command(label="Copy to Clipboard", command=lambda: copy_to_clipboard(window, listbox))
  share_menu.add_command(label="Share via Email", command=lambda: share_via_email(window, listbox))
  

def create_window():
  global status_label
  global history_manager
  history_manager = HistoryManager()
  window.title("Number List")
  version_label = tk.Label(window, text="No version available")
  version_label.pack()
  global listbox
  undo_redo_manager = UndoRedoManager()
  undo_button = tk.Button(window, text="Undo", command=lambda: undo(undo_redo_manager, listbox))
  undo_button.pack()
  redo_button = tk.Button(window, text="Redo", command=lambda: redo(undo_redo_manager, listbox))
  redo_button.pack()
  delete_button = tk.Button(window,
                            text="Delete Selected Entry",
                            command=lambda: delete_selected_entry(listbox, undo_redo_manager))
  delete_button.pack()  
  menubar = tk.Menu(window)
  window.config(menu=menubar)
  file_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="File", menu=file_menu)
  file_menu.add_command(label="New", command=lambda: create_new_window())
  file_menu.add_command(label="Import",
                        command=lambda: open_file(window, listbox))
  file_menu.add_command(label="Exit", command=lambda: exit_file(window))
  export_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Export", menu=export_menu)
  export_menu.add_command(label="Export to CSV", command=lambda: export_to_csv(window, listbox))
  export_menu.add_command(label="Export to JSON", command=lambda: export_to_json(window, listbox))
  export_menu.add_command(label="Export to Excel", command=lambda: export_to_excel(window, listbox))
  export_menu.add_command(label="Export to ODS", command=lambda: export_to_ods(window, listbox))
  edit_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Edit", menu=edit_menu)
  edit_menu.add_command(label="Undo", command=lambda: undo(undo_redo_manager, listbox))
  edit_menu.add_command(label="Redo", command=lambda: redo(undo_redo_manager, listbox))
  edit_menu.add_command(label="Remove Duplicates", command=lambda: remove_duplicates(listbox))
  help_menu = tk.Menu(menubar, tearoff=0)
  help_menu.add_command(label="Report a Bug", command=lambda: report_bug(window))
  menubar.add_cascade(label="Help", menu=help_menu)
  help_menu.add_command(label="About", command=lambda: about(window))
  math_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Calculate", menu=math_menu)
  data_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="Data", menu=data_menu)
  data_menu.add_command(label="Numeral System Conversions", command=lambda: numeral_system_conversions(listbox, history_manager))
  math_menu.add_command(label="Add",
                        command=lambda: add_all_numbers(window, listbox, history_manager, version_label))
  math_menu.add_command(label="Subtract",
                        command=lambda: subtract_numbers(window, listbox, history_manager, version_label))
  math_menu.add_command(label="Multiply",
                        command=lambda: multiply_all_numbers(window, listbox, history_manager, version_label))
  math_menu.add_command(label="Divide",
                        command=lambda: divide_all_numbers(window, listbox, history_manager, version_label))
  math_menu.add_command(label="Square",
                        command=lambda: square_all_numbers(window, listbox, history_manager, version_label))
  math_menu.add_command(label="Define Algebraic Letter",
                        command=lambda: define_algebraic_letter(window))
  more_algebra_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="More Algebra...", menu=more_algebra_menu)
  more_algebra_menu.add_command(
    label="Convert Algebra", command=lambda: convert_algebra(window, listbox))
  sort_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="Sort", menu=sort_menu)
  sort_menu.add_command(label="Ascending", command=lambda: sort_numbers_ascending(window, listbox, undo_redo_manager, history_manager))
  sort_menu.add_command(label="Descending", command=lambda: sort_numbers_descending(window, listbox, undo_redo_manager, history_manager))
  filter_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="Filter", menu=filter_menu)
  filter_menu.add_command(label="Even Numbers", 
                            command=lambda: filter_even_numbers(window, listbox, undo_redo_manager, history_manager))
  filter_menu.add_command(label="Odd Numbers", 
                            command=lambda: filter_odd_numbers(window, listbox, undo_redo_manager, history_manager))
  filter_menu.add_command(label="Custom Range", 
                            command=lambda: filter_custom_range(window, listbox, undo_redo_manager, history_manager))
  graph_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Graph", menu=graph_menu)
  graph_menu.add_command(label="Create Graph", command=lambda: create_graph(window, listbox))
  graph_menu.add_command(label="Create Advanced Graph", command=lambda: create_advanced_graph(window, listbox))
  graph_menu.add_command(label="Create Histogram", command=lambda: create_histogram(window, listbox))
  graph_menu.add_command(label="Create Box Plot", command=lambda: create_box_plot(window, listbox))
  history_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="History", menu=history_menu)
  history_menu.add_command(label="View History", command=lambda: view_history(history_manager))
  history_menu.add_command(label="Undo to Previous State", command=lambda: restore_history(history_manager.undo()))
  history_menu.add_command(label="Redo to Next State", command=lambda: restore_history(history_manager.redo()))
  history_menu.add_command(label="Save Named Version", command=lambda: save_named_version(window, listbox, history_manager))
  stats_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Statistics", menu=stats_menu)
  stats_menu.add_command(label="Mean", command=lambda: calculate_mean(listbox))
  stats_menu.add_command(label="Median", command=lambda: calculate_median(listbox))
  stats_menu.add_command(label="Mode", command=lambda: calculate_mode(listbox))
  stats_menu.add_command(label="Variance", command=lambda: calculate_variance(listbox))
  stats_menu.add_command(label="Standard Deviation", command=lambda: calculate_standard_deviation(listbox))
  transform_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Transform", menu=transform_menu)
  normalize_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Normalize", menu=normalize_menu)
  normalize_menu.add_command(label="Min-Max Scaling", 
                               command=lambda: create_transformation_window(window, listbox, 'min_max'))
  normalize_menu.add_command(label="Z-Score Normalization", 
                               command=lambda: create_transformation_window(window, listbox, 'z_score'))
  scale_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Scale", menu=scale_menu)
  scale_menu.add_command(label="Logarithmic Scaling", 
                           command=lambda: create_transformation_window(window, listbox, 'log'))
  scale_menu.add_command(label="Exponential Scaling", 
                           command=lambda: create_transformation_window(window, listbox, 'exp'))
  bin_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Binning", menu=bin_menu)
  bin_menu.add_command(label="Equal-width Binning", 
                         command=lambda: create_transformation_window(window, listbox, 'equal_width_bin'))
  bin_menu.add_command(label="Equal-frequency Binning", 
                         command=lambda: create_transformation_window(window, listbox, 'equal_freq_bin'))
  outlier_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Outlier Handling", menu=outlier_menu)
  outlier_menu.add_command(label="Remove Outliers", 
                             command=lambda: create_transformation_window(window, listbox, 'remove_outliers'))
  outlier_menu.add_command(label="Cap Outliers", 
                             command=lambda: create_transformation_window(window, listbox, 'cap_outliers'))
  transform_menu.add_command(label="Impute Missing Values", 
                               command=lambda: create_transformation_window(window, listbox, 'impute_missing'))
  time_series_menu = tk.Menu(transform_menu, tearoff=0)
  transform_menu.add_cascade(label="Time Series", menu=time_series_menu)
  time_series_menu.add_command(label="Create Lag", 
                                 command=lambda: create_transformation_window(window, listbox, 'lag'))
  time_series_menu.add_command(label="Difference", 
                                 command=lambda: create_transformation_window(window, listbox, 'difference'))
  time_series_menu.add_command(label="Rolling Statistics", 
                                 command=lambda: create_transformation_window(window, listbox, 'rolling'))
  listbox = tk.Listbox(window)
  listbox.pack()
  entry = tk.Entry(window)
  entry.pack()
  button_add = tk.Button(window, text="Add number", 
                         command=lambda: add_number(window, listbox, entry, undo_redo_manager, history_manager))
  button_add.pack()
  button_clear = tk.Button(window,
                           text="Clear List",
                           command=lambda: clear_list(listbox, history_manager))
  button_clear.pack()
  status_label = create_status_bar(window)
  session_manager.load_session(listbox)
  update_status(status_label, f"List loaded. Current size: {listbox.size()}")
  bind_keyboard_shortcuts(window, listbox, undo_redo_manager, entry)
  theme_menu = tk.Menu(menubar, tearoff=0)
  theme_menu.add_command(label="Light Theme", command=lambda: change_theme("light"))
  theme_menu.add_command(label="Dark Theme", command=lambda: change_theme("dark"))
  menubar.add_cascade(label="Themes", menu=theme_menu)
  share_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Share", menu=share_menu)
  share_menu.add_command(label="Copy to Clipboard", command=lambda: copy_to_clipboard(window, listbox))
  share_menu.add_command(label="Share via Email", command=lambda: share_via_email(window, listbox))
create_window() 
session_manager.load_session(listbox)

def main():
  window.mainloop()

main()