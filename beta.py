import tkinter as tk
import csv
from datetime import datetime
from tkinter import messagebox, filedialog
from tkinter import simpledialog
import openpyxl
import xlwt
import xlrd
import pandas as pd
import ezodf
from ezodf import Sheet
from odf.opendocument import OpenDocumentText
from odf.text import P

window = tk.Tk()
listbox = tk.Listbox(window)
counter = 1

def delete_selected_entry(listbox):
  selected_index = listbox.curselection()
  if selected_index:
    listbox.delete(selected_index)
  else:
    messagebox.showerror("Error", "No entry selected!", parent=window)

def show_current_file_extension(window):
  messagebox.showinfo(
    "Current File Extension",
    f"The current file extension is: {window.file_extension}")

def ask_file_type(window):
  file_type = simpledialog.askstring(
    "Input",
    "Enter the file type (.csv, .xls, .odt, .ods, .xlsx)",
    parent=window)

  window.file_extension = "." + file_type.split(".")[-1]

  return file_type

file_extension = ask_file_type(window)

def save_list(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error", "You can't save an empty list!", parent=window)
    return

  try:
    file_type = window.file_extension
    if file_type == '.csv':
      save_to_csv(window, listbox)
    elif file_type == '.xls':
      save_to_xls(window, listbox)
    elif file_type == '.odt':
      save_to_odf(window, listbox)
    elif file_type == '.ods':
      save_to_ods(window, listbox)
    elif file_type == '.xlsx':
      save_to_xlsx(window, listbox)
    else:
      messagebox.showerror("Error", "Invalid file type!", parent=window)

  except Exception as e:
    messagebox.showerror("Error", str(e), parent=window)

def add_number(window, listbox, entry):
  global counter
  item = entry.get()
  if not item:
    messagebox.showerror("Error",
                         "You can't add an empty entry!",
                         parent=window)
    return
  if item.isdigit() or float or item in algebra_dict:
    listbox.insert(tk.END, f"{counter}. {item}")
    counter += 1
  else:
    messagebox.showerror(
      "Error",
      "You can only add numbers or defined algebraic letters!",
      parent=window)
  entry.delete(0, tk.END)

def clear_list(listbox, show_message=True):
  numbers = listbox.get(0, tk.END)
  if not numbers and show_message:
    messagebox.showerror("Error", "The list is already empty!")
  else:
    listbox.delete(0, tk.END)

from odf import text, teletype
from odf.opendocument import load

def open_file(window, listbox):
  global counter
  filename = filedialog.askopenfilename(filetypes=[
    ('Excel Files', '*.xls ; *.xlsx'), ('CSV Files', '*.csv'),
    ('ODF Text Files', '*.odt'), ('ODF Spreadsheet Files', '*.ods'),
    ('All Files', '*.*')
  ], parent=window)
  if not filename:
    return
  clear_list(listbox, show_message=False)
  window.file_extension = "." + filename.split(".")[-1]
  if filename.endswith('.xls'):
    book = xlrd.open_workbook(filename)
    sheet = book.sheet_by_index(0)
    for i in range(sheet.nrows):
      listbox.insert(tk.END, f"{counter}. {str(int(sheet.cell_value(i, 0)))}")
      counter += 1
  elif filename.endswith('.csv'):
    with open(filename, 'r') as f:
      reader = csv.reader(f)
      for row in reader:
        for number in row:
          listbox.insert(tk.END, f"{counter}. {number}")
          counter += 1
  elif filename.endswith('.odt'):
    textdoc = load(filename)
    allparas = textdoc.getElementsByType(text.P)
    for para in allparas:
      number = teletype.extractText(para)
      listbox.insert(tk.END, f"{counter}. {number}")
      counter += 1
  elif filename.endswith('.ods'):
    spreadsheet = ezodf.opendoc(filename).sheets[0]
    for row in spreadsheet.rows():
      for cell in row:
        if cell.value is not None:
          listbox.insert(tk.END, f"{counter}. {str(cell.value)}")
  elif filename.endswith('.xlsx'):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_cols(min_row=1, min_col=1, values_only=True):
      for cell in row:
        if cell is not None:
          listbox.insert(tk.END, f"{counter}. {str(cell)}")
          counter += 1

def save_to_csv(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error",
                         "You can't save an empty list!",
                         parent=window)
    return
  now = datetime.now()
  timestamp = now.strftime("%Y%m%d%H%M%S")
  filename = f'numbers_{timestamp}.csv'
  with open(filename, 'w', newline='') as f:
    writer = csv.writer(f)
    for number in numbers:
      writer.writerow([number.split(". ")[1]])

def save_to_ods(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error",
                         "You can't save an empty list!",
                         parent=window)
    return
  now = datetime.now()
  timestamp = now.strftime("%Y%m%d%H%M%S")
  filename = f'numbers_{timestamp}.ods'
  spreadsheet = ezodf.newdoc(doctype='ods', filename=filename)
  sheet = Sheet('Sheet1', size=(len(numbers), 1))
  for i, number in enumerate(numbers):
    sheet[i, 0].set_value(number.split(". ")[1])
  spreadsheet.sheets += sheet
  spreadsheet.save()

def save_to_xls(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error",
                         "You can't save an empty list!",
                         parent=window)
    return
  messagebox.showinfo(
    "XLS Info",
    "Replit has a strange bug where if you view the XLS file in the file explorer, download said file, and then choose to open it in Excel, it will corrupt the file. You will have to download the file without opening it through Replit to properly view the file, unfortunately this cannot be remedied at this time.",
    parent=window)
  now = datetime.now()
  timestamp = now.strftime("%Y%m%d%H%M%S")
  filename = f'numbers_{timestamp}.xls'
  book = xlwt.Workbook()
  sheet1 = book.add_sheet('Sheet 1')
  for i, number in enumerate(numbers):
    sheet1.write(i, 0, number.split(". ")[1])
  book.save(filename)
  if not filename:
    return
  if filename.endswith('.csv'):
    with open(filename, 'r') as f:
      reader = csv.reader(f)
      clear_list(listbox)
      for row in reader:
        for number in row:
          listbox.insert(tk.END, f"{counter}. {number}")
          counter += 1
  elif filename.endswith('.xls'):
    df = pd.read_excel(filename)
    clear_list(listbox)
    for index, row in df.iterrows():
      for number in row:
        listbox.insert(tk.END, f"{counter}. {number}")
        counter += 1
  else:
    messagebox.showerror("Error", "Unsupported file format.", parent=window)

def save_to_xlsx(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error",
                         "You can't save an empty list!",
                         parent=window)
    return
  now = datetime.now()
  filename = f'numbers_{now.strftime("%Y%m%d%H%M%S")}.xlsx'
  filetypes = [('Excel Files', ['*.xlsx ; *.xls'])]
  filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                          filetypes=filetypes,
                                          parent=window)
  if not filename:
    return
  workbook = openpyxl.Workbook()
  sheet = workbook.active
  for i, number in enumerate(numbers):
    sheet.cell(row=i + 1, column=1, value=number.split(". ")[1])
  workbook.save(filename).xlsx

def file_menu_change(file_menu):
  file_menu.add_command(label="Save as XLSX",
                        command=lambda: save_to_xlsx(window, listbox))

def save_to_odf(window, listbox):
  numbers = listbox.get(0, tk.END)
  if not numbers:
    messagebox.showerror("Error",
                         "You can't save an empty list!",
                         parent=window)
    return
  now = datetime.now()
  timestamp = now.strftime("%Y%m%d%H%M%S")
  filename = f'numbers_{timestamp}.odt'
  textdoc = OpenDocumentText()
  for number in numbers:
    p = P(text=str(number.split(". ")[1]))
    textdoc.text.addElement(p)
  textdoc.save(filename)

def add_all_numbers(window, listbox):
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 2:
    messagebox.showerror("Error",
                         "There must be numbers to add!",
                         parent=window)
    return
  total = sum(int(number.split(". ")[1]) for number in numbers)
  clear_list(listbox)
  listbox.insert(tk.END, f"{counter}. {total}")

def subtract_numbers(window, listbox):
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 2:
    messagebox.showerror("Error",
                         "There must be numbers to subtract!",
                         parent=window)
    return
  total = int(numbers[0].split(". ")[1]) - sum(
    int(number.split(". ")[1]) for number in numbers[1:])
  clear_list(listbox)
  listbox.insert(tk.END, f"{counter}. {total}")

def multiply_all_numbers(window, listbox):
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 2:
    messagebox.showerror("Error",
                         "There must be numbers to multiply!",
                         parent=window)
    return
  total = 1
  for number in numbers:
    total *= int(number.split(". ")[1])
  clear_list(listbox)
  listbox.insert(tk.END, f"{counter}. {total}")

def divide_all_numbers(window, listbox):
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 2:
    messagebox.showerror("Error",
                         "There must be at least two numbers to divide!",
                         parent=window)
    return
  total = int(numbers[0].split(". ")[1])
  for number in numbers[1:]:
    if int(number.split(". ")[1]) == 0:
      messagebox.showerror("Error", "Cannot divide by zero!", parent=window)
      return
    total /= int(number.split(". ")[1])
  clear_list(listbox)
  listbox.insert(tk.END, f"{counter}. {total}")

def square_all_numbers(window, listbox):
  numbers = listbox.get(0, tk.END)
  if len(numbers) < 1:
    messagebox.showerror("Error",
                         "There must be at least one number to square!",
                         parent=window)
    return
  clear_list(listbox)
  for number in numbers:
    square_number = float(number.split(". ")[1])**2
    listbox.insert(tk.END, f"{counter}. {square_number}")

algebra_dict = {}

def define_algebraic_letter(window):
  letter_window = tk.Toplevel(window)
  letter_window.title("Define Algebraic Letter")
  tk.Label(letter_window, text="Algebraic Letter:").pack()
  letter_entry = tk.Entry(letter_window)
  letter_entry.pack()
  tk.Label(letter_window, text="Equivalent Number:").pack()
  number_entry = tk.Entry(letter_window)
  number_entry.pack()

  def save_algebraic_letter():
    global algebra_dict
    algebra_dict[letter_entry.get()] = int(number_entry.get())
    letter_window.destroy()

  tk.Button(letter_window, text="Save", command=save_algebraic_letter).pack()

def convert_algebra(window, listbox):
  counter = 1
  items = listbox.get(0, tk.END)
  converted_items = []
  for item in items:
    if item in algebra_dict:
      converted_items.append(str(algebra_dict[item]))
    else:
      converted_items.append(item)
    listbox.delete(0, tk.END)
  for item in converted_items:
    listbox.insert(tk.END, f"{counter}. {item}")
    counter += 1
  if set(items) == set(converted_items):
    messagebox.showerror("Error",
                         "You can't convert non-defined algebraic numbers!",
                         parent=window)

def change_file_extension(window, extension):
  window.file_extension = extension

about_window = None

def about(window):
  global about_window
  if about_window is not None:
    try:
      about_window.lift()
    except tk.TclError:
      about_window = None
  if about_window is None:
    about_window = tk.Toplevel(window)
    about_window.protocol("WM_DELETE_WINDOW", close_about)
  title_label = tk.Label(about_window, text="About Number List:")
  title_label.pack()
  update_label = tk.Label(about_window, text="The 'We Love More Files' Update")
  update_label.pack()
  version_label = tk.Label(about_window, text="Version 0.62.887 BETA")
  version_label.pack()
  contributor_label = tk.Label(about_window, text="Contributors:")
  contributor_label.pack()
  contributor_label2 = tk.Label(about_window, text="Tay Rake 2023")
  contributor_label2.pack()

def close_about():
  global about_window
  about_window.destroy()

about_window = None
  
def exit_file(window):
    window.quit()

def create_new_window():
  window = tk.Tk()
  window.title("Number List")
  delete_button = tk.Button(window,
                            text="Delete Selected Entry",
                            command=lambda: delete_selected_entry(listbox))
  delete_button.pack()
  button_extension = tk.Button(window,
                               text="Change File Extension",
                               command=lambda: ask_file_type(window))
  button_extension.pack()
  menubar = tk.Menu(window)
  window.config(menu=menubar)
  file_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="File", menu=file_menu)
  file_menu.add_command(label="Save",
                        command=lambda: save_list(window, listbox))
  file_menu.add_command(label="New", command=lambda: create_new_window())
  file_menu.add_command(label="Open",
                        command=lambda: open_file(window, listbox))
  file_menu.add_command(label="Exit", command=lambda: exit_file(window))
  help_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Help", menu=help_menu)
  help_menu.add_command(label="About", command=lambda: about(window))
  help_menu.add_command(label="Current File Extension",
                        command=lambda: show_current_file_extension(window))
  math_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Math", menu=math_menu)
  math_menu.add_command(label="Add",
                        command=lambda: add_all_numbers(window, listbox))
  math_menu.add_command(label="Subtract",
                        command=lambda: subtract_numbers(window, listbox))
  math_menu.add_command(label="Multiply",
                        command=lambda: multiply_all_numbers(window, listbox))
  math_menu.add_command(label="Divide",
                        command=lambda: divide_all_numbers(window, listbox))
  math_menu.add_command(label="Square",
                        command=lambda: square_all_numbers(window, listbox))
  math_menu.add_command(label="Define Algebraic Letter",
                        command=lambda: define_algebraic_letter(window))
  more_algebra_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="More Algebra...", menu=more_algebra_menu)
  more_algebra_menu.add_command(
    label="Convert Algebra", command=lambda: convert_algebra(window, listbox))
  listbox = tk.Listbox(window)
  listbox.pack()
  entry = tk.Entry(window)
  entry.pack()
  button_add = tk.Button(window,
                         text="Add number",
                         command=lambda: add_number(window, listbox, entry))
  button_add.pack()
  button_clear = tk.Button(window,
                           text="Clear List",
                           command=lambda: clear_list(listbox))
  button_clear.pack()

def create_window():
  window.title("Number List")
  delete_button = tk.Button(window,
                            text="Delete Selected Entry",
                            command=lambda: delete_selected_entry(listbox))
  delete_button.pack()
  button_extension = tk.Button(window,
                               text="Change File Extension",
                               command=lambda: ask_file_type(window))
  button_extension.pack()
  menubar = tk.Menu(window)
  window.config(menu=menubar)
  file_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="File", menu=file_menu)
  file_menu.add_command(label="Save",
                        command=lambda: save_list(window, listbox))
  file_menu.add_command(label="New", command=lambda: create_new_window())
  file_menu.add_command(label="Open",
                        command=lambda: open_file(window, listbox))
  file_menu.add_command(label="Exit", command=lambda: exit_file(window))
  help_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Help", menu=help_menu)
  help_menu.add_command(label="About", command=lambda: about(window))
  help_menu.add_command(label="Current File Extension",
                        command=lambda: show_current_file_extension(window))
  math_menu = tk.Menu(menubar, tearoff=0)
  menubar.add_cascade(label="Math", menu=math_menu)
  math_menu.add_command(label="Add",
                        command=lambda: add_all_numbers(window, listbox))
  math_menu.add_command(label="Subtract",
                        command=lambda: subtract_numbers(window, listbox))
  math_menu.add_command(label="Multiply",
                        command=lambda: multiply_all_numbers(window, listbox))
  math_menu.add_command(label="Divide",
                        command=lambda: divide_all_numbers(window, listbox))
  math_menu.add_command(label="Square",
                        command=lambda: square_all_numbers(window, listbox))
  math_menu.add_command(label="Define Algebraic Letter",
                        command=lambda: define_algebraic_letter(window))
  more_algebra_menu = tk.Menu(math_menu, tearoff=0)
  math_menu.add_cascade(label="More Algebra...", menu=more_algebra_menu)
  more_algebra_menu.add_command(
    label="Convert Algebra", command=lambda: convert_algebra(window, listbox))
  listbox = tk.Listbox(window)
  listbox.pack()
  entry = tk.Entry(window)
  entry.pack()
  button_add = tk.Button(window,
                         text="Add number",
                         command=lambda: add_number(window, listbox, entry))
  button_add.pack()
  button_clear = tk.Button(window,
                           text="Clear List",
                           command=lambda: clear_list(listbox))
  button_clear.pack()

create_window()
messagebox.showinfo(
  "Saving Info",
  "Defining file extensions have changed again! Starting from Version 0.60, clicking on the 'Change File Extension' button will now change the file extension of the file you are currently working on! No more navigating to the 'File' menu!"
)
messagebox.showinfo(
  "Saving Info #2",
  "Saving files in general have also changed! Starting from Version 0.59, saving lists can now only be done through the 'File' submenu. Happy saving!"
)

def main():
  window.mainloop()

main()
